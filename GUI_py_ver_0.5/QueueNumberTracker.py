import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from PIL import Image, ImageTk
import pyautogui
import cv2
import numpy as np
import yaml
import time
import pytesseract
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import webbrowser
import re
import os
import sys
import threading
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO

# 2025/8/14
# 語言設定
lang = "zh"  #"en" 或 "zh"

''' 更新英文版預定 '''

# 載入 YAML 資料
with open("hospital_doctors.yaml", "r", encoding="utf-8") as f:
    hospital_data = yaml.safe_load(f)

# 將來的 Excel 先在全域宣告
file_path = None  

# 建立主視窗
root = tk.Tk()
root.title("螢幕擷取控制 GUI")
root.geometry("720x480")

stop_event = threading.Event()  # 全局停止旗標

# 第三頁關閉視窗
def on_closing():
    stop_event.set()
    # 判斷是按鈕關閉還是視窗關閉可以用參數或拆成兩個函式
    # 這裡假設是視窗關閉事件，直接關閉視窗
    root.destroy()

# 第三頁停止執行 => go_to_page4
def on_close_button():
    stop_event.set()
    go_to_page4()

# 攔截視窗關閉事件
root.protocol("WM_DELETE_WINDOW", on_closing)

''' -------------------- 涵式區 -------------------- '''
# 函數：在文字框中新增輸出
def append_output(text):
    text_output.config(state="normal")      # 解除唯讀
    text_output.insert("end", text + "\n")  # 插入新文字並加上換行
    text_output.see("end")                  # 自動捲到最底
    text_output.config(state="disabled")    # 保持唯讀狀態

def capture_screen(region):
    """
    擷取螢幕特定範圍畫面。
    
    Args:
        region (tuple): 範圍 (x, y, width, height)
    Returns:
        numpy.ndarray: 擷取的影像
    """
    screenshot = pyautogui.screenshot(region=region)
    image = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2BGR)
    
    # 儲存圖片，參數：(名稱, 變數) cv2.imwrite('test.jpg', image)

    return image

def recognize_digits(image, return_binary_image = 0):
    """
    辨識影像中的數字。
    
    Args:
        image (numpy.ndarray): 影像數據
    Returns:
        str: 辨識結果
    """
    # 將影像轉為灰階
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # 使用二值化處理
    _, binary_image = cv2.threshold(gray, 128, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    # 黑白顛倒(變成白底黑字)
    binary_image = cv2.bitwise_not(binary_image)
  
    # 使用 Tesseract OCR 辨識數字；oem：辨識引擎、psm：輸入版面格式，只辨識數字（0-9），忽略文字和符號
    custom_config = r'--oem 3 --psm 6 outputbase digits'
    result = pytesseract.image_to_string(binary_image, config=custom_config)
    
    if return_binary_image:
        return binary_image, result
    
    else:
        return result

def write_to_excel(latest_number, now_time, file_path):
    """
    將兩個參數寫入 Excel 檔案的最下一行。
    
    Args:
        latest_number (int): 要寫入的數字。
        now_time (str): 當下時間，格式可以是字符串。
        file_path (str): Excel 檔案的路徑。
    """
    try:
        # 嘗試加載現有 Excel 文件
        workbook = load_workbook(file_path)
        sheet = workbook.active  # 使用默認的第一個工作表
        print(f'成功連接：{file_path}')
        append_output(f'成功連接：{file_path}')

        latest_number = latest_number.strip()   # 刪除字串頭尾的換行和空白字元
        sheet.append([latest_number, now_time]) # 找到最後一行，在其後新增一行
        
        # 保存 Excel 文件
        workbook.save(file_path)
        print(f"寫入 {file_path}: {latest_number}, {now_time}") 
        append_output(f"寫入辨識結果: {latest_number}, 時間戳記： {now_time}") 

    except Exception as e:
        # 捕捉其他所有例外，並印出錯誤訊息
        print(f'發生錯誤: {e}')
        append_output(f'發生錯誤: {e}')
        return 1    # 觸發主程式 break

    return 0

# 檢查目前是否打包成 exe 並切換路徑
def get_save_folder():
    if getattr(sys, 'frozen', False):  # 判斷是否為打包模式
        base_path = os.path.dirname(sys.executable)  # exe 所在資料夾
    else:
        base_path = os.path.abspath(".")
    
    save_folder = os.path.join(base_path, "data")

    # 如果 data 資料夾不存在，就建立它
    if not os.path.exists(save_folder):
        os.makedirs(save_folder)

    return save_folder

# 更新第一頁顯示畫面
def update_capture():
    # 擷取螢幕範圍（使用 pyautogui 擷取）
    x, y = capture_params['x'], capture_params['y']
    w, h = capture_params['width'], capture_params['height']
    frame = capture_screen((x, y, w, h))
    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)  # 轉回 Tkinter 可用格式

    ''' 將 frame 縮放成固定大小 (選取範圍時不要壓縮圖片比較好)
    frame = cv2.resize(frame, (DISPLAY_WIDTH, DISPLAY_HEIGHT))  
    '''

    img = Image.fromarray(frame)
    imgtk = ImageTk.PhotoImage(image=img)

    page1_video_label.imgtk = imgtk
    page1_video_label.configure(image=imgtk)

    # 每 100ms 更新一次
    page1_video_label.after(100, update_capture)

# 刷新醫生選單(包含網址)
def update_doctor_menu(event=None):
    hospital_info = hospital_data.get(selected_hospital.get(), {})
    doctors = hospital_info.get("doctors", [])
    url = hospital_info.get("url", "（None）")
    
    doctor_menu["values"] = doctors
    if doctors:
        selected_doctor.set(doctors[0])
    else:
        selected_doctor.set("None")

    hospital_url_var.set(f"{url}")

# URL 超連結
def open_url(event):
    url = hospital_url_var.get()
    if url:
        webbrowser.open(url)

# 滑鼠移入：改變樣式（加底線、改顏色）
def on_enter(event):
    url_label_2.config(font=("Arial", 12, "underline"), fg="darkblue")

# 滑鼠移出：恢復原來樣式
def on_leave(event):
    url_label_2.config(font=("Arial", 12), fg="blue")
 
# 滑桿變更事件
def slider_changed(param, val):
    capture_params[param] = int(val)
    entry_vars[param].set(str(val))  # 同步 Entry 顯示

# 檢查第一頁圖大小是否大於第二頁顯示空間，如果有，等比例壓縮
def Check_image_size(width, height, pil_img):
    if width > 360 or height > 240:
        scale_w = 360 / width
        scale_h = 240 / height
        scale = min(scale_w, scale_h)  # 取較小比例，確保不超過限制
        new_width = int(width * scale)
        new_height = int(height * scale)
        pil_img = pil_img.resize((new_width, new_height), resample=Image.Resampling.LANCZOS)
    
    return pil_img

# 建立 Excel，回傳 file_path
def build_Execl(timer_val, file_name):
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    h = selected_hospital.get()
    d = selected_doctor.get()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Dr.{d}"           # 工作表名稱
    sheet.append([date, f"{timer_val} h", f"{h}"]) # 日期、時長、醫院
    sheet.append(["號次", "時間", "間隔"]) 

    # 對 A1:C1 置中
    for row in sheet["A1:C2"]:  # 兩列都置中
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 保存 Excel 文件
    try:
        # 將 Excel 存在資料夾 data
        save_folder = get_save_folder()
        file_path = os.path.join(save_folder, f"{file_name}.xlsx")

        workbook.save(file_path)
        print(f"已建立檔案於: {file_path}")
        append_output(f"已建立檔案於: {file_path}")

        return file_path

    except Exception as e:
        print(f"建立檔案失敗，原因：\n{e}")
        append_output(f"建立檔案失敗，原因：\n{e}")

# 資料處裡演算法
def Data_processing_algorithms(file_path):
    # 使用第一個分頁作為來源
    wb = load_workbook(file_path)
    source_ws = wb.worksheets[0]

    # 從第 3 列開始加公式(第一頁)
    for row in range(3, source_ws.max_row):
        formula = f"=(B{row+1}-B{row})*86400"
        source_ws.cell(row=row, column=3).value = formula

    # 建立新的工作表，若已存在名為「校正版」的工作表就刪掉（避免重複）
    if "校正版" in wb.sheetnames:
        del wb["校正版"]
    target_ws = wb.create_sheet("校正版")

    # 先複製標頭（前兩列）
    for row_index in [1, 2]:
        row_data = [source_ws.cell(row=row_index, column=col).value for col in range(1, 8)]
        target_ws.append(row_data)
    
    # 校正資料
    max_row = source_ws.max_row
    last_number = -1            # 上一筆號次
    DPM = 0                     # 偏移量
    threshold_of_refresh = 15   # 刷新錯誤時間閥值
    threshold_of_Pass = 65      # 過號時間閥值

    ''' 校正演算法 + 生成校正資料 '''
    for row in range(3, max_row + 1):
        row_data = []  # 每列要寫入的資料

        # 號次
        try:
            value = int(source_ws.cell(row=row, column=1).value)
        except:
            print(f"第 {row} 列異常：號次 = {source_ws.cell(row=row, column=1).value}，視為雜訊忽略不計")
            DPM += 1
            continue

        # 號次異常條件檢查
        if value > 200 or value < 0 or value == last_number:
            DPM += 1
            continue

        row_data.append(value)

        # 時間
        try:
            time_value = source_ws.cell(row=row, column=2).value
            next_time_value = source_ws.cell(row=row+1, column=2).value
            time_1 = datetime.strptime(time_value, "%H:%M:%S")
            time_2 = datetime.strptime(next_time_value, "%H:%M:%S")

        except:
            if  row == max_row:
                time_1 = time_2
            else:
                print("時間數據欄位出現格式錯誤")

        time_diff = int((time_2 - time_1).total_seconds())    # 計算時間差

        # 檢查是否小於等於 threshold_of_refresh 秒，(最後一筆資料時間差=0)
        if 0 < time_diff and time_diff <= threshold_of_refresh:
            print(f"第 {row} 列異常：時間差 {time_diff} 秒，視為雜訊忽略不計")
            DPM += 1
            continue
        
        # (演算邏輯)確定沒有刷新錯誤後才更新參數:"上一筆數字"
        last_number = value
        row_data.append(time_value)

        # 間隔（公式，注意要修正偏移）
        formula = f"=(B{(row - DPM) + 1}-B{(row - DPM)})*86400"
        row_data.append(formula)

        # 以 row 為單位加入
        target_ws.append(row_data)

    ''' <<< 統計數據 >>> '''

    # 儲存檔案以讀取計算後數據
    wb.save(file_path)
    wb = load_workbook(file_path)
    target_ws = wb["校正版"]
    max_row = target_ws.max_row

    valid_data = 1 # 有效數據編號
    chart2_y_data = []

    for row in range(4, max_row):
            
        time_value = target_ws.cell(row=row, column=2).value
        next_time_value = target_ws.cell(row=row+1, column=2).value
        time_1 = datetime.strptime(time_value, "%H:%M:%S")
        time_2 = datetime.strptime(next_time_value, "%H:%M:%S")
        time_diff = int((time_2 - time_1).total_seconds())    # 計算時間差

        # 檢查是否小於等於 threshold_of_refresh 秒，(最後一筆資料時間差=0)
        if time_diff > threshold_of_Pass:
            target_ws[f'D{row}'] = valid_data
            valid_data += 1
            chart2_y_data.append(time_diff)

    target_ws['D2'] = '有效編號'
    target_ws['F1'] = '資料數量'
    target_ws['G1'] = '有效數量'
    target_ws['H1'] = '過濾數量'

    target_ws['F2'] = f'=COUNTA(A3:A{max_row})'                              # 資料數量
    target_ws['G2'] = f'=COUNTIFS(C4:C{max_row-1}, ">={threshold_of_Pass}")' # 有效資料數量
    target_ws['H2'] = '=F2-G2'

    target_ws['F5'] = '平均'
    target_ws['G5'] = '標準差'
    target_ws['H5'] = '變異係數'

    target_ws['F6'] = f'=AVERAGE(C4:C{max_row-1})'  # 平均值
    target_ws['G6'] = f'=STDEV(C4:C{max_row-1})'    # 標準差
    target_ws['H6'] = '=F6/F2'                      # 變異系數

    # 設定置中對齊
    center_alignment = Alignment(horizontal='center', vertical='center')

    for row in target_ws.iter_rows():
        for cell in row:
            if cell.value is not None:  # 只對有資料的儲存格套用對齊
                cell.alignment = center_alignment

    # 提示語
    target_ws['F4'] = f"過濾 頭尾 & 過號 (間隔小於 {threshold_of_Pass} s) 的數據所計算的結果" 
    target_ws.append([])
    target_ws.append([f'以上為程式自動校正，過濾不正常 & 間隔小於 {threshold_of_refresh} s 的號次'])

    user_choice = image_option.get()
    ''' <<< 使用 matplotlib 繪圖 >>> '''
    if user_choice == "yes":

        plt.rcParams['font.family'] = 'Microsoft JhengHei'  # 設定中文字型
        plt.rcParams['axes.unicode_minus'] = False          # 避免負號顯示錯誤（常見中文+數字混合問題）

        ''' 圖片一：號次與時間'''
        # 擷取資料（x:號碼, y:時間）
        x_data, y_data= [], []

        for row in target_ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=2):
            x_val = row[0].value
            y_val = row[1].value
            if x_val and y_val:
                x_data.append(x_val)
                y_data.append(y_val)

        plt.figure(figsize=(10, 6))
        plt.plot(x_data, y_data, marker='o', linestyle='-')
        plt.title("關係圖：號次與時間", fontsize=20)
        plt.xlabel("號次", fontsize=16)
        plt.ylabel("時間", fontsize=16)
        plt.grid(True)
        plt.gcf().autofmt_xdate()   # 旋轉 X 軸標籤避免擠在一起

        # 儲存成圖片(在 memory)
        buf1 = BytesIO()
        plt.savefig(buf1, format='png', bbox_inches='tight')
        plt.close()

        # 插入圖片到 Excel
        buf1.seek(0)    # 將 檔案指標 移到開頭
        img1 = ExcelImage(buf1)
        img1.anchor = "F8"
        target_ws.add_image(img1)

        ''' 圖片二：有效編號與間隔'''
        # 擷取資料（x:號碼, y:時間）
        x_data = []

        # 擷取資料：X 軸為 D 欄，Y 軸為 C 欄，排除 D 欄為空
        for row in target_ws.iter_rows(min_row=3, max_row=max_row, min_col=3, max_col=4):
            x_val = row[1].value  # D欄
            if x_val is not None:
                x_data.append(x_val)

        plt.figure(figsize=(10, 6))
        plt.plot(x_data, chart2_y_data, marker='o', linestyle='-')
        plt.title("關係圖：有效編號與間隔", fontsize=20)
        plt.xlabel("有效編號", fontsize=16)
        plt.ylabel("間隔", fontsize=16)
        plt.grid(True)

        buf2 = BytesIO()
        plt.savefig(buf2, format='png', bbox_inches='tight')
        plt.close()
        buf2.seek(0)
        img2 = ExcelImage(buf2)
        img2.anchor = "T8"
        target_ws.add_image(img2)

    # 儲存檔案
    wb.save(file_path)
    print(f"已完成檔案 {file_path} 校正。")

''' --------------------< 頁面跳轉函式 >-------------------- '''
def go_to_page1():
    page2_frame.pack_forget()                   # pack 暫時不顯示，設定還在
    page1_frame.pack(fill="both", expand=True)  # 填滿畫面、容器空間變大時跟著撐開

def go_to_page2():
    # 從 capture_params 讀取參數
    x = capture_params['x']
    y = capture_params['y']
    width = capture_params['width']
    height = capture_params['height']
    region = (x, y, width, height)      # 螢幕擷取範圍 (x, y, width, height)
    captured_image = capture_screen(region)
    binary_image, digits = recognize_digits(captured_image, 1)

    # 將 binary_image 轉成適合Tkinter顯示的格式
    pil_img = Image.fromarray(binary_image)
    
    # 檢查第一頁圖大小是否大於第二頁顯示空間，如果有，等比例壓縮
    pil_img = Check_image_size(width, height, pil_img)

    # PIL Image 再轉 PhotoImage
    photo_img = ImageTk.PhotoImage(pil_img)

    binary_img_canvas.delete("all")     # 清除舊圖
    binary_img_canvas.create_image(180, 120, image=photo_img, anchor="center")  # 座標非長寬
    binary_img_canvas.image = photo_img  # 保持引用避免被 Python 自動垃圾回收機制回收

    # 辨識結果垂直滾動條
    digit_result_text.delete("1.0", tk.END)    # 刷新前需要先清空     
    digit_result_text.insert(tk.END, f"辨識結果：{digits}")

    h = selected_hospital.get()
    d = selected_doctor.get()
    page2_check_label.config(text=f"資料出處\n{h}\n{d} 醫師")

    # 計算檔案名稱（日期 + 部門）
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    department = h.split("_")[1]
    file_name = f"{date}_{department}_{d}"
    file_name_var.set(file_name)    # 輸入框預設值

    page1_frame.pack_forget()
    page2_frame.pack(fill="both", expand=True)

def go_to_page3():
    global file_path

    # 取得第二頁輸入值確保其合法
    file_name = file_name_var.get()
    timer_val = Timer_input.get()
    interval_val = interval_input.get()

    # 檔名檢查
    if not file_name.strip():
        messagebox.showerror("錯誤", "檔名不能為空")
        return

    # 不允許的字元 (Windows 常見不合法字元)
    invalid_chars = r'[\\/:*?"<>|]'
    if re.search(invalid_chars, file_name):
        messagebox.showerror("錯誤", "檔名包含不允許的字元: \\ / : * ? \" < > |")
        return
    
    # 檢查檔案是否已存在
    save_folder = get_save_folder()  
    file_full_path = os.path.join(save_folder, f"{file_name}.xlsx")  
    if os.path.exists(file_full_path):
        messagebox.showerror("錯誤", f"檔案 '{file_name}.xlsx' 已存在，請更換檔名")
        return
    
    # 檢查是否為正整數且符合範圍
    if not timer_val.isdigit() or not (0 < int(timer_val) <= 10):
        messagebox.showerror("錯誤", "執行時間必須為 1~10 小時之間的正整數")
        Timer_entry.focus_set()
        return  

    if not interval_val.isdigit() or not (0 < int(interval_val) <= 600):
        messagebox.showerror("錯誤", "辨識間隔必須為 1~600 秒之間的正整數")
        interval_entry.focus_set()
        return  
    
    # 建立 Excel   
    file_path = build_Execl(timer_val, file_name)

    page2_frame.pack_forget()
    page3_frame.pack(fill="both", expand=True)

    ''' --------------------< 以下辨識迴圈(副執行序) >-------------------- '''
    def recognition_loop():
        timer_value = int(Timer_input.get())
        interval_value = int(interval_input.get())
        total_seconds = timer_value * 3600
        start_time = time.time()
        region = (capture_params['x'], capture_params['y'], capture_params['width'], capture_params['height'])
        latest_number = -1

        while  not stop_event.is_set():
            elapsed = time.time() - start_time
            if elapsed > total_seconds:
                append_output(">>> 已達設定執行時間，結束辨識")
                break

            captured_image = capture_screen(region)
            digits = recognize_digits(captured_image)

            print(f"辨識結果: {digits}")
            append_output(f"辨識結果: {digits}")

            if digits != latest_number:
                latest_number = digits
                now = datetime.now()
                now_time = now.strftime("%H:%M:%S")
                if write_to_excel(latest_number, now_time, file_path):
                    break

            # 分段睡眠，方便快速響應關閉事件
            sleep_time = 0
            while sleep_time < interval_value:
                if stop_event.is_set():
                    break
                time.sleep(0.2)
                sleep_time += 0.2

    # 加開執行緒給辨識迴圈，以免 sleep 導致 GUI 卡死
    thread = threading.Thread(target=recognition_loop, daemon=True)
    thread.start()  

    ''' --------------------< 以上辨識迴圈 >-------------------- '''

def go_to_page4():
    page3_frame.pack_forget()
    page4_frame.pack(fill="both", expand=True)

def go_to_page5():
    try:
        Data_processing_algorithms(file_path)
    except Exception as e:
        print(f"發生錯誤：{str(e)}")
        label5_text.set(f"發生錯誤：\n{str(e)}")

    page4_frame.pack_forget()
    page5_frame.pack(fill="both", expand=True)

''' ---------------------------------------- 以下 GUI ---------------------------------------- '''


''' 選取範圍時不要壓縮圖片比較好
DISPLAY_WIDTH = 360   # 固定顯示區寬
DISPLAY_HEIGHT = 240  # 固定顯示區高
'''

''' 
--------------------<<< 第一頁 >>>-------------------- 
'''

page1_frame = tk.Frame(root)
page1_frame.pack(fill="both", expand=True)

''' ----------< 以下選單 >---------- '''
hospital_names = list(hospital_data.keys())

# 建立一個子 frame 來放在同一列的元件
row_frame = tk.Frame(page1_frame)
row_frame.pack(pady=(20,0))

selected_hospital = tk.StringVar()          # 醫院選單變數
selected_hospital.set(hospital_names[0])    # 預設第一個
selected_doctor = tk.StringVar()            # 醫生選單變數
hospital_url_var = tk.StringVar()           # 網址變數

# 醫院 Label、選單
hospital_label = tk.Label(row_frame, text="選擇醫院:", font=("Arial", 18))
hospital_label.grid(row=0, column=0, padx=5)
hospital_menu = ttk.Combobox(row_frame, textvariable=selected_hospital, values=hospital_names, state="readonly", font=("Arial", 18), width=21)
hospital_menu.grid(row=0, column=1, padx=10)

# 醫生 Label、選單
doctor_label = tk.Label(row_frame, text="選擇醫生:", font=("Arial", 18))
doctor_label.grid(row=0, column=2, padx=5)
doctor_menu = ttk.Combobox(row_frame, textvariable=selected_doctor, state="readonly", font=("Arial", 18), width=7)
doctor_menu.grid(row=0, column=3, padx=10)

# 網址顯示 Label（新增）醫院網址: 
url_label_1 = tk.Label(row_frame, text="醫院網址:", font=("Arial", 18))
url_label_1.grid(row=1, column=0, padx=5)
url_label_2 = tk.Label(row_frame, textvariable=hospital_url_var, font=("Arial", 12), fg="blue", anchor="w", justify="left", cursor="hand2", wraplength=570)
url_label_2.grid(row=1, column=1, columnspan=4, sticky="w", pady=(0, 0))

# 超連結按鈕(綁定滑鼠事件)
url_label_2.bind("<Button-1>", open_url)    # 點擊
url_label_2.bind("<Enter>", on_enter)       # 滑入
url_label_2.bind("<Leave>", on_leave)       # 滑出

# 選擇醫院時重新呼叫 update_doctor_menu
hospital_menu.bind("<<ComboboxSelected>>", update_doctor_menu)

# 初始化醫生選單
update_doctor_menu()

''' ----------< 以上選單 >----------'''

# 左側顯示畫面
page1_video_label = tk.Label(page1_frame)
page1_video_label.pack(side=tk.LEFT, padx=(20,0))

# 右側控制區
control_frame = tk.Frame(page1_frame)
control_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(0,20))

''' ----------< 以下拉桿 >----------'''

# 建立四個拉桿與標籤
sliders = {}
value_labels = {}
slider_range = {
    'x': (0, 1920),
    'y': (0, 1080),
    'width': (200, 480),
    'height': (200, 360)
}

# 拉桿範圍參數
capture_params = {
    'x': 0,
    'y': 0,
    'width': 480,
    'height': 360
}

sliders = {}
entry_vars = {}  # 儲存文字輸入框的變數

for idx, (param, (min_val, max_val)) in enumerate(slider_range.items()):
    row_base = idx * 2  # 每個參數占兩行：一行 Entry，一行滑桿

    # Label設定
    tk.Label(control_frame, text=f"{param}：", width=7, anchor="center").grid(
        row=row_base, column=0, sticky="w", padx=5, pady=(20, 7))

    # 輸入框設定
    var = tk.StringVar()
    var.set(str(capture_params[param]))
    entry = tk.Entry(control_frame, textvariable=var, width=8, justify="center")
    entry.grid(row=row_base, column=1, sticky="w", pady=(20, 7))

    # 滑桿設定
    slider = tk.Scale(
        control_frame,
        from_=min_val,
        to=max_val,
        orient=tk.HORIZONTAL,
        command=lambda val, p=param: slider_changed(p, val),
        showvalue=False,  # 隱藏上方數值
        length=150
    )

    slider.set(capture_params[param])
    slider.grid(row=row_base + 1, column=0, columnspan=2, sticky="we", padx=5)

    # Entry 改變時同步滑桿
    def entry_callback(event, p=param, v=var, s=slider):
        try:
            new_val = int(v.get())
            new_val = max(min(new_val, slider_range[p][1]), slider_range[p][0])
            capture_params[p] = new_val
            s.set(new_val)
        except ValueError:
            v.set(str(capture_params[p]))  # 無效輸入時還原

    entry.bind("<Return>", entry_callback)
    entry.bind("<FocusOut>", entry_callback)

    # 儲存元件
    sliders[param] = slider
    entry_vars[param] = var

''' ----------< 以上拉桿 >---------- '''

# 確定按鈕
btn_confirm = tk.Button(control_frame,text="確定", command=go_to_page2, font=("Arial", 18, "bold"), padx=10, pady=5, relief="raised", bg="#F44336", fg="white", activebackground="#D32F2F", bd=5)
btn_confirm.place(relx=1.0, rely=1.0, anchor="se", x=0, y=-10)

'''
--------------------<<< 第二頁 >>>--------------------
'''
page2_frame = tk.Frame(root)
label2 = tk.Label(page2_frame, text="                             確認資料                             ", font=("Arial", 20), bg="#F44336")
label2.pack(pady=(20,0))

# 左側顯示畫面
page2_left_frame = tk.Frame(page2_frame)
page2_left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(20,0))
page2_left_up = tk.Frame(page2_left_frame)
page2_left_up.pack(pady=10, anchor="nw")
page2_left_down = tk.Frame(page2_left_frame)
page2_left_down.pack(fill="both", pady=(0,10))

# 右側控制區
page2_right_frame = tk.Frame(page2_frame)
page2_right_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(0,20))
page2_right_up = tk.Frame(page2_right_frame)
page2_right_up.pack(pady=10, anchor="ne")    # 靠右上排列

''' ----------< 以下 第二頁 左側 >---------- '''
image_text_label = tk.Label(page2_left_up, text="處理後影像", font=("Arial", 20))
image_text_label.pack(anchor="center")

# 處理後畫面
binary_img_canvas = tk.Canvas(page2_left_up, width=360, height=240, bg="white", highlightthickness=0)
binary_img_canvas.pack(side=tk.TOP)

# 辨識結果(滾動條)
digit_result_text = tk.Text(page2_left_down, font=("Arial", 18), height=1, width=23, wrap="word")
scrollbar = tk.Scrollbar(page2_left_down, command=digit_result_text.yview)
digit_result_text.configure(yscrollcommand=scrollbar.set)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
digit_result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

''' ----------< 以下 第二頁 右側 >---------- '''
# 顯示檔名
tk.Label(page2_right_up, text="建立檔案", font=("Arial", 20)).pack()

# 檔名輸入框
file_name_var = tk.StringVar(value="file_name")
file_name_entry_frame = tk.Frame(page2_right_up)
file_name_entry_frame.pack(anchor="e", pady=(5,0))

tk.Entry(page2_right_up, textvariable=file_name_var, font=("Arial", 15), width=20).pack(side="left")
tk.Label(page2_right_up, text=".xlsx", font=("Arial", 20)).pack(side="left")

# 顯示選擇的醫院與醫生
page2_check_label = tk.Label(page2_right_frame, text="", font=("Arial", 20)) # text 更新於涵式：update_page2_content()
page2_check_label.pack(pady=20)

# 計時器輸入框
Timer_frame = tk.Frame(page2_right_frame)
Timer_frame.pack(padx=(5, 10))
tk.Label(Timer_frame, text="執行時間：", font=("Arial", 20)).pack(side="left")
Timer_input = tk.StringVar(value="3")    # 執行時間預設3小時
Timer_entry = tk.Entry(Timer_frame, textvariable=Timer_input, font=("Arial", 20), width=3, justify="right")
Timer_entry.pack(side="left", padx=(5, 10))
tk.Label(Timer_frame, text="h", font=("Arial", 20)).pack(side="left")

# 間隔輸入框
interval_frame = tk.Frame(page2_right_frame)
interval_frame.pack(padx=(5, 10))
tk.Label(interval_frame, text="辨識間隔：", font=("Arial", 20)).pack(side="left")
interval_input = tk.StringVar(value="10")    # 間隔時間預設10s
interval_entry = tk.Entry(interval_frame, textvariable=interval_input, font=("Arial", 20), width=3, justify="right")
interval_entry.pack(side="left", padx=(5, 10))
tk.Label(interval_frame, text="s", font=("Arial", 20)).pack(side="left")

# 開始執行、返回按鈕
btn_start = tk.Button(page2_frame, text="開始辨識", command=go_to_page3, font=("Arial", 18, "bold"), padx=10, pady=5, relief="raised", bg="#F44336", fg="white", activebackground="#D32F2F", bd=5)
btn_start.place(relx=1.0, rely=1.0, anchor="se", x=-20, y=-10)
btn_back = tk.Button(page2_frame, text="返回", command=go_to_page1, font=("Arial", 18, "bold"), padx=10, pady=5, relief="raised", bg="#9E9E9E", fg="white", activebackground="#7E7E7E", bd=5)
btn_back.place(relx=0.0, rely=1.0, anchor="sw", x=20, y=-10)

'''
--------------------<<< 第三頁 >>>--------------------
新增第三頁輸出時，使用涵式 append_output(value), type:'str'
'''
page3_frame = tk.Frame(root)

label3 = tk.Label(page3_frame, text="系統運作中", font=("Arial", 24))
label3.pack(pady=(20,10))

''' ----------< 以下輸出面板 >----------'''
'''
規格：
    wrap="word":               換行時以「單字」為單位，而不是硬切字元。
    bg="black", fg="white":    文字框背景黑色、字體白色，模仿終端機樣式。
    insertbackground="white":  設定游標顏色為白色（輸入模式時可見）。
    font=("Consolas", 12):     採用等寬字體 Consolas,12 號字，方便對齊。
    state="disabled":          預設禁用輸入，避免使用者手動改內容。
    pack(side="left", fill="both", expand=True)：放在左邊並填滿空間
'''

output_frame = tk.Frame(page3_frame, bg="black", bd=2, relief="sunken")
output_frame.pack(padx=40, pady=(0,80), fill="both", expand=True)

text_output = tk.Text(output_frame, wrap="word", bg="black", fg="white", insertbackground="white", font=("Consolas", 12), state="disabled")
text_output.pack(side="left", fill="both", expand=True)

# 滾動條
scrollbar = tk.Scrollbar(output_frame, command=text_output.yview)
scrollbar.pack(side="right", fill="y")
text_output.config(yscrollcommand=scrollbar.set)

''' ----------< 以上輸出面板 >----------'''

# 按鈕
btn_stop_page3 = tk.Button(page3_frame, text="中止執行", command=on_close_button, font=("Arial", 18, "bold"), padx=10, pady=5, relief="groove", bg="#F44336", fg="white", activebackground="#D32F2F", bd=7)
btn_stop_page3.place(relx=0.5, rely=1.0, anchor="s", y=-10)

'''
--------------------<<< 第四頁 >>>--------------------
'''
page4_frame = tk.Frame(root)

label4 = tk.Label(page4_frame, text="影像辨識結束", font=("Arial", 32))
label4.pack(fill="x", pady=(20,10))

# 顯示數據處理設定
settings_label = tk.Label(page4_frame, text="設定數據處裡閥值參數:", font=("Arial", 24), anchor="w")
settings_label.pack(fill="x", padx=50, pady=(10,0))

# 閥值一(網頁刷新錯誤) 輸入框
Threshold_1_frame = tk.Frame(page4_frame)
Threshold_1_frame.pack(pady=10)
tk.Label(Threshold_1_frame, text="(排除網頁刷新)閥值一：", font=("Arial", 24)).pack(side="left")
Threshold_1_input = tk.StringVar(value="15")    # 預設 15 s
Threshold_1_entry = tk.Entry(Threshold_1_frame, textvariable=Threshold_1_input, font=("Arial", 24), width=3, justify='center')
Threshold_1_entry.pack(side="left", padx=(5, 10))
tk.Label(Threshold_1_frame, text="s", font=("Arial", 24)).pack(side="left")

# 閥值二輸入框
Threshold_2_frame = tk.Frame(page4_frame)
Threshold_2_frame.pack()
tk.Label(Threshold_2_frame, text="(排除過號號次)閥值二：", font=("Arial", 24)).pack(side="left")
Threshold_2_input = tk.StringVar(value="65")    # 預設 65 s
Threshold_2_entry = tk.Entry(Threshold_2_frame, textvariable=Threshold_2_input, font=("Arial", 24), width=3, justify='center')
Threshold_2_entry.pack(side="left", padx=(5, 10))
tk.Label(Threshold_2_frame, text="s", font=("Arial", 20)).pack(side="left")

# 生成圖片? (選項)
option_label = tk.Label(page4_frame, text="產生圖片", font=("Arial", 24))
option_label.pack(pady=(20, 5))

image_option = tk.StringVar(value="yes")  # 預設 yes
option_frame = tk.Frame(page4_frame)
option_frame.pack()

yes_btn = tk.Radiobutton(option_frame, text="Yes", variable=image_option, value="yes", font=("Arial", 20))
yes_btn.pack(side="left", padx=5)
no_btn = tk.Radiobutton(option_frame, text="No", variable=image_option, value="no", font=("Arial", 20))
no_btn.pack(side="left", padx=5)

# 按鈕 
btn_start_page4 = tk.Button(page4_frame, text="開始處理", command=go_to_page5, font=("Arial", 18, "bold"), padx=10, pady=5, relief="raised", bg="#F44336", fg="white", activebackground="#D32F2F", bd=7)
btn_start_page4.place(relx=1.0, rely=1.0, anchor="se", x=-150, y=-10)
btn_stop_page4 = tk.Button(page4_frame, text="結束程式", command=root.destroy, font=("Arial", 18, "bold"), padx=10, pady=5, relief="raised", bg="#9E9E9E", fg="white", activebackground="#D32F2F", bd=7)
btn_stop_page4.place(relx=0.0, rely=1.0, anchor="sw", x=150, y=-10)
'''
--------------------<<< 第五頁 >>>--------------------
'''
page5_frame = tk.Frame(root)

label5_text = tk.StringVar()
label5_text.set("已生成校正後數據") 

label5 = tk.Label(page5_frame, textvariable=label5_text, font=("Arial", 32), wraplength=600, justify="left")
label5.pack(pady=150)

btn_stop_page5 = tk.Button(page5_frame, text="結束程式", command=root.destroy, font=("Arial", 18, "bold"), padx=10, pady=5, relief="groove", bg="#F44336", fg="white", activebackground="#D32F2F", bd=7)
btn_stop_page5.place(relx=0.5, rely=1.0, anchor="s", y=-10)

# 開始畫面更新
update_capture()

root.mainloop()